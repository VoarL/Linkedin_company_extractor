"""
Job Descriptions Extractor

This script reads the Job Tracker Excel file, opens each LinkedIn job link,
extracts the full job description, and writes them to text files organized by category.

Each category (Digital, Analog, etc.) gets its own text file with format:
    Company: <company name>
    Job Title: <job title>
    URL: <link>

    <job description>

    ----------------------------------------
"""

import os
import re
import time
import random
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime


def setup_driver():
    """Set up Chrome driver with options to avoid detection."""
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(options=chrome_options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def is_linkedin_job_url(url):
    """Check if the URL is a valid LinkedIn job URL."""
    if not url or not isinstance(url, str):
        return False

    linkedin_patterns = [
        r'linkedin\.com/jobs/view/\d+',
        r'linkedin\.com/jobs/search.*currentJobId=\d+',
    ]

    return any(re.search(pattern, url) for pattern in linkedin_patterns)


def get_url_from_cell(cell):
    """Extract URL from cell - either from hyperlink or cell value."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    if cell.value and isinstance(cell.value, str):
        return cell.value
    return None


def format_description_from_html(html):
    """Convert HTML to plain text while preserving bullet points."""
    import re

    # Replace <li> tags with bullet points
    html = re.sub(r'<li[^>]*>', '\n  • ', html)
    html = re.sub(r'</li>', '', html)

    # Replace <br> tags with newlines
    html = re.sub(r'<br\s*/?>', '\n', html)

    # Replace paragraph/div closings with double newlines
    html = re.sub(r'</p>', '\n\n', html)
    html = re.sub(r'</div>', '\n', html)

    # Replace headers with newlines
    html = re.sub(r'</(h[1-6])>', '\n\n', html)

    # Remove all remaining HTML tags
    html = re.sub(r'<[^>]+>', '', html)

    # Decode HTML entities
    html = html.replace('&nbsp;', ' ')
    html = html.replace('&amp;', '&')
    html = html.replace('&lt;', '<')
    html = html.replace('&gt;', '>')
    html = html.replace('&quot;', '"')
    html = html.replace('&#39;', "'")

    # Clean up excessive whitespace
    lines = html.split('\n')
    cleaned_lines = []
    prev_empty = False

    for line in lines:
        line = line.strip()
        if line:
            cleaned_lines.append(line)
            prev_empty = False
        elif not prev_empty:
            cleaned_lines.append('')
            prev_empty = True

    return '\n'.join(cleaned_lines).strip()


def get_job_site_type(url):
    """Identify the job board type from URL."""
    if not url:
        return None
    if 'linkedin.com' in url:
        return 'linkedin'
    if 'greenhouse.io' in url:
        return 'greenhouse'
    if 'myworkdayjobs.com' in url or 'workday.com' in url:
        return 'workday'
    if 'lever.co' in url:
        return 'lever'
    if 'oraclecloud.com' in url:
        return 'oracle'
    if 'hrmdirect.com' in url:
        return 'hrmdirect'
    if 'careers.' in url or 'jobs.' in url:
        return 'generic'
    return 'generic'


def extract_generic_job_info(driver, url):
    """Extract job info from generic job pages using common patterns."""
    try:
        driver.get(url)
        time.sleep(random.uniform(3, 5))

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        company = None
        job_title = None
        description = None

        # Common title selectors across job sites
        title_selectors = [
            'h1[class*="title"]', 'h1[class*="job"]', 'h1[class*="posting"]',
            'h1[data-automation*="title"]', '.job-title', '.posting-title',
            'h1', 'h2[class*="title"]'
        ]

        for selector in title_selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element and element.text.strip():
                    job_title = element.text.strip()
                    break
            except NoSuchElementException:
                continue

        # Common company selectors
        company_selectors = [
            '[class*="company"]', '[data-automation*="company"]',
            '.employer-name', '.company-name', 'a[href*="/company"]'
        ]

        for selector in company_selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element and element.text.strip():
                    company = element.text.strip()
                    break
            except NoSuchElementException:
                continue

        # Common description selectors
        desc_selectors = [
            '[class*="description"]', '[class*="job-content"]',
            '[data-automation*="description"]', '.job-details',
            '[class*="posting-content"]', '[class*="job-body"]',
            'article', '.content'
        ]

        for selector in desc_selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element:
                    html = element.get_attribute('innerHTML')
                    if html and len(html) > 100:
                        description = format_description_from_html(html)
                        if description and len(description) > 50:
                            break
            except NoSuchElementException:
                continue

        return {
            'company': company,
            'job_title': job_title,
            'description': description
        }

    except TimeoutException:
        print(f"    Timeout loading page")
        return None
    except Exception as e:
        print(f"    Error: {str(e)}")
        return None


def extract_greenhouse_job_info(driver, url):
    """Extract job info from Greenhouse job pages."""
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 4))

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        company = None
        job_title = None
        description = None

        # Greenhouse-specific selectors
        try:
            element = driver.find_element(By.CSS_SELECTOR, 'h1.app-title, h1[class*="title"]')
            job_title = element.text.strip() if element else None
        except NoSuchElementException:
            pass

        try:
            element = driver.find_element(By.CSS_SELECTOR, '.company-name, [class*="company"]')
            company = element.text.strip() if element else None
        except NoSuchElementException:
            pass

        # Job description
        try:
            element = driver.find_element(By.CSS_SELECTOR, '#content, .content, [class*="description"]')
            if element:
                html = element.get_attribute('innerHTML')
                description = format_description_from_html(html) if html else None
        except NoSuchElementException:
            pass

        return {
            'company': company,
            'job_title': job_title,
            'description': description
        }

    except Exception as e:
        print(f"    Greenhouse error: {str(e)}")
        return extract_generic_job_info(driver, url)


def extract_workday_job_info(driver, url):
    """Extract job info from Workday job pages."""
    try:
        driver.get(url)
        time.sleep(random.uniform(3, 5))

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        company = None
        job_title = None
        description = None

        # Workday-specific selectors
        try:
            element = driver.find_element(By.CSS_SELECTOR, '[data-automation-id="jobPostingHeader"], h2[data-automation-id="jobTitle"], h1')
            job_title = element.text.strip() if element else None
        except NoSuchElementException:
            pass

        # Description - Workday often has it in a specific div
        try:
            element = driver.find_element(By.CSS_SELECTOR, '[data-automation-id="jobPostingDescription"], [class*="jobDescription"]')
            if element:
                html = element.get_attribute('innerHTML')
                description = format_description_from_html(html) if html else None
        except NoSuchElementException:
            pass

        if not description:
            return extract_generic_job_info(driver, url)

        return {
            'company': company,
            'job_title': job_title,
            'description': description
        }

    except Exception as e:
        print(f"    Workday error: {str(e)}")
        return extract_generic_job_info(driver, url)


def extract_hrmdirect_job_info(driver, url):
    """Extract job info from HRMDirect job pages."""
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 4))

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        company = None
        job_title = None
        description = None

        # HRMDirect-specific selectors
        try:
            element = driver.find_element(By.CSS_SELECTOR, '.careersTitle, h1')
            job_title = element.text.strip() if element else None
        except NoSuchElementException:
            pass

        # Job description - HRMDirect uses .jobDesc
        try:
            element = driver.find_element(By.CSS_SELECTOR, '.jobDesc, div.jobDesc')
            if element:
                html = element.get_attribute('innerHTML')
                description = format_description_from_html(html) if html else None
        except NoSuchElementException:
            pass

        # If .jobDesc didn't work, try getting all content after the title
        if not description:
            try:
                element = driver.find_element(By.CSS_SELECTOR, '.reqResult, #content, body')
                if element:
                    html = element.get_attribute('innerHTML')
                    description = format_description_from_html(html) if html else None
            except NoSuchElementException:
                pass

        return {
            'company': company,
            'job_title': job_title,
            'description': description
        }

    except Exception as e:
        print(f"    HRMDirect error: {str(e)}")
        return extract_generic_job_info(driver, url)


def extract_job_info_any(driver, url):
    """Extract job info from any supported job site."""
    site_type = get_job_site_type(url)
    print(f"    Site type: {site_type}")

    if site_type == 'linkedin':
        return extract_job_info(driver, url)
    elif site_type == 'greenhouse':
        return extract_greenhouse_job_info(driver, url)
    elif site_type == 'workday':
        return extract_workday_job_info(driver, url)
    elif site_type == 'hrmdirect':
        return extract_hrmdirect_job_info(driver, url)
    else:
        return extract_generic_job_info(driver, url)


def extract_job_info(driver, url):
    """Extract job title, company, and description from a LinkedIn job page."""
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 4))

        company = None
        job_title = None
        description = None

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # Extract job title
        for selector in ['h1.top-card-layout__title', 'h1.topcard__title', 'h1']:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element and element.text.strip():
                    job_title = element.text.strip()
                    break
            except NoSuchElementException:
                continue

        # Extract company name
        for selector in ['a.topcard__org-name-link', '.topcard__flavor a', 'a[href*="/company/"]']:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element and element.text.strip():
                    company = element.text.strip()
                    break
            except NoSuchElementException:
                continue

        # Try to click "Show more" button to expand description
        try:
            show_more_btn = driver.find_element(By.CSS_SELECTOR, '.show-more-less-html__button--more')
            show_more_btn.click()
            time.sleep(0.5)
        except:
            pass  # Button may not exist or already expanded

        # Extract job description with bullet points preserved
        description_selectors = [
            '.show-more-less-html__markup',
            '.description__text',
            '.jobs-description__content',
            '.jobs-box__html-content',
            'div[class*="description"]',
            '.job-details',
        ]

        for selector in description_selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element:
                    # Get inner HTML and convert to text with bullets preserved
                    html = element.get_attribute('innerHTML')
                    if html:
                        description = format_description_from_html(html)
                        if description:
                            break
            except NoSuchElementException:
                continue

        return {
            'company': company,
            'job_title': job_title,
            'description': description
        }

    except TimeoutException:
        print(f"    Timeout loading page")
        return None
    except Exception as e:
        print(f"    Error: {str(e)}")
        return None


def sanitize_filename(name):
    """Remove invalid characters from filename."""
    return re.sub(r'[<>:"/\\|?*]', '', name)


def normalize_url(url):
    """Normalize LinkedIn URL for comparison (remove tracking params)."""
    if not url:
        return None
    # Extract the base job ID from LinkedIn URLs
    match = re.search(r'linkedin\.com/jobs/view/(\d+)', url)
    if match:
        return f"linkedin.com/jobs/view/{match.group(1)}"
    return url.split('?')[0].rstrip('/')


def get_existing_urls(filename):
    """Parse an existing txt file and return dict of URLs with their status.

    Returns:
        dict with 'extracted' (successfully extracted URLs) and 'skipped' (URLs marked as SKIPPED)
    """
    result = {'extracted': set(), 'skipped': set()}
    if not os.path.exists(filename):
        return result

    try:
        with open(filename, 'r', encoding='utf-8') as f:
            content = f.read()

            # Split by job entries (separated by dashes)
            entries = re.split(r'-{40,}', content)

            for entry in entries:
                # Find URL in this entry
                url_match = re.search(r'URL: (https?://[^\s\n]+)', entry)
                if not url_match:
                    continue

                url = url_match.group(1)
                normalized = normalize_url(url)
                if not normalized:
                    continue

                # Check if this entry was skipped
                if 'Status: SKIPPED' in entry or 'Status: ERROR' in entry:
                    result['skipped'].add(normalized)
                else:
                    result['extracted'].add(normalized)

    except Exception as e:
        print(f"Warning: Could not read existing file {filename}: {e}")

    return result


def main():
    excel_path = "Job Tracker.xlsx"

    print(f"Reading {excel_path}...")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find column indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[header] = col

    print(f"Headers found: {headers}")

    category_col = headers.get('Category', 1)
    company_col = headers.get('Company', 2)
    job_title_col = headers.get('Job Title', 3)  # This column has hyperlinks to jobs

    # Organize jobs by category
    jobs_by_category = {}
    errors = []

    for row in range(2, ws.max_row + 1):
        category = ws.cell(row=row, column=category_col).value
        if not category:
            continue

        category = category.strip()
        if category not in jobs_by_category:
            jobs_by_category[category] = []

        # Get URL from Job Title column hyperlink
        job_title_cell = ws.cell(row=row, column=job_title_col)
        url = get_url_from_cell(job_title_cell)

        # Get existing company/title from spreadsheet as fallback
        existing_company = ws.cell(row=row, column=company_col).value
        existing_title = job_title_cell.value

        jobs_by_category[category].append({
            'row': row,
            'url': url,
            'existing_company': existing_company,
            'existing_title': existing_title,
            'is_linkedin': is_linkedin_job_url(url) if url else False
        })

    # Count jobs
    total_jobs = sum(len(jobs) for jobs in jobs_by_category.values())
    linkedin_jobs = sum(1 for jobs in jobs_by_category.values() for job in jobs if job['is_linkedin'])

    print(f"\nFound {total_jobs} jobs across {len(jobs_by_category)} categories")
    print(f"LinkedIn jobs: {linkedin_jobs}")
    print(f"Non-LinkedIn/No URL: {total_jobs - linkedin_jobs}")
    print(f"\nCategories: {list(jobs_by_category.keys())}")

    # Check existing files and filter out already-extracted jobs
    jobs_to_extract = {}
    total_new_jobs = 0
    total_already_done = 0
    total_retry_skipped = 0

    for category, jobs in jobs_by_category.items():
        filename = f"{sanitize_filename(category)}_jobs.txt"
        url_status = get_existing_urls(filename)
        extracted_urls = url_status['extracted']
        skipped_urls = url_status['skipped']

        new_jobs = []
        for job in jobs:
            url = job['url']
            normalized = normalize_url(url) if url else None

            if normalized and normalized in extracted_urls:
                # Already successfully extracted
                total_already_done += 1
            elif normalized and normalized in skipped_urls:
                # Was skipped before - retry it
                new_jobs.append(job)
                total_retry_skipped += 1
                total_new_jobs += 1
            else:
                # Brand new job
                new_jobs.append(job)
                total_new_jobs += 1

        jobs_to_extract[category] = {
            'jobs': new_jobs,
            'filename': filename,
            'file_exists': os.path.exists(filename),
            'existing_count': len(extracted_urls),
            'skipped_count': len(skipped_urls)
        }

        if len(extracted_urls) > 0 or len(skipped_urls) > 0:
            print(f"  {category}: {len(extracted_urls)} done, {len(skipped_urls)} skipped (retrying), {len(new_jobs)} to process")

    print(f"\nTotal: {total_already_done} jobs already extracted, {total_retry_skipped} previously skipped (retrying), {total_new_jobs} to process")

    if total_new_jobs == 0:
        print("\nNo new jobs to extract. All jobs are already in the text files.")
        return

    # Set up browser
    print("\nSetting up browser...")
    driver = setup_driver()

    try:
        processed = 0
        for category, data in jobs_to_extract.items():
            jobs = data['jobs']
            filename = data['filename']

            if not jobs:
                print(f"\n{category}: No new jobs to process")
                continue

            print(f"\n{'='*60}")
            print(f"Processing category: {category} ({len(jobs)} new jobs)")
            print('='*60)

            # Append to existing file or create new one
            mode = 'a' if data['file_exists'] else 'w'

            with open(filename, mode, encoding='utf-8') as f:
                # Write header only for new files
                if not data['file_exists']:
                    f.write(f"{'='*60}\n")
                    f.write(f"Category: {category}\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"{'='*60}\n\n")
                else:
                    # Add a separator for appended content
                    f.write(f"\n{'='*60}\n")
                    f.write(f"Appended: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"{'='*60}\n\n")

                for i, job in enumerate(jobs):
                    processed += 1
                    row = job['row']
                    url = job['url']

                    print(f"\n[{processed}/{total_new_jobs}] Row {row}...")

                    if not url:
                        # No URL at all
                        error_msg = f"Row {row}: No URL"
                        errors.append(error_msg)
                        print(f"  Skipping: {error_msg}")

                        f.write(f"Company: {job['existing_company'] or 'N/A'}\n")
                        f.write(f"Job Title: {job['existing_title'] or 'N/A'}\n")
                        f.write(f"URL: N/A\n")
                        f.write(f"Status: SKIPPED - No URL\n")
                        f.write(f"\n{'-'*40}\n\n")
                        continue

                    print(f"  URL: {url[:50]}...")
                    info = extract_job_info_any(driver, url)

                    if info:
                        company = info['company'] or job['existing_company'] or 'N/A'
                        title = info['job_title'] or job['existing_title'] or 'N/A'
                        description = info['description'] or 'No description available'

                        print(f"  Company: {company}")
                        print(f"  Title: {title}")
                        print(f"  Description: {len(description)} chars")

                        f.write(f"Company: {company}\n")
                        f.write(f"Job Title: {title}\n")
                        f.write(f"URL: {url}\n")
                        f.write(f"\n{description}\n")
                        f.write(f"\n{'-'*40}\n\n")
                    else:
                        error_msg = f"Row {row}: Failed to extract - {url[:50]}..."
                        errors.append(error_msg)
                        print(f"  ERROR: Failed to extract")

                        f.write(f"Company: {job['existing_company'] or 'N/A'}\n")
                        f.write(f"Job Title: {job['existing_title'] or 'N/A'}\n")
                        f.write(f"URL: {url}\n")
                        f.write(f"Status: ERROR - Failed to extract description\n")
                        f.write(f"\n{'-'*40}\n\n")

                    # Random delay
                    time.sleep(random.uniform(2, 4))

            print(f"\nSaved: {filename}")

    finally:
        driver.quit()

    # Write error log
    if errors:
        with open('extraction_errors.txt', 'w', encoding='utf-8') as f:
            f.write(f"Extraction Errors - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*60 + "\n\n")
            for error in errors:
                f.write(f"{error}\n")
        print(f"\n{len(errors)} errors logged to extraction_errors.txt")

    print(f"\nDone! Processed {processed} jobs.")
    print(f"Output files created for categories: {list(jobs_by_category.keys())}")


if __name__ == "__main__":
    main()
