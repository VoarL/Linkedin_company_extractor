"""
LinkedIn Job Details Extractor

This script reads an Excel file, extracts job details from LinkedIn URLs in column B,
and fills columns D (Company), E (Job Title), F (How long ago) with the extracted data.

It handles both:
- Plain URLs in cells
- Hyperlinks where the display text might be different from the URL
"""

import re
import time
import random
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException


def setup_driver():
    """Set up Chrome driver with options to avoid detection."""
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(options=chrome_options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def is_linkedin_job_url(url):
    """Check if the URL is a valid LinkedIn job URL."""
    if not url or not isinstance(url, str):
        return False

    linkedin_patterns = [
        r'linkedin\.com/jobs/view/\d+',
        r'linkedin\.com/jobs/search.*currentJobId=\d+',
    ]

    return any(re.search(pattern, url) for pattern in linkedin_patterns)


def get_url_from_cell(cell):
    """Extract URL from cell - either from hyperlink or cell value."""
    # First check if there's a hyperlink
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target

    # Otherwise check the cell value
    if cell.value and isinstance(cell.value, str):
        return cell.value

    return None


def extract_job_details(driver, url):
    """Extract job details from a LinkedIn job page."""
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 4))  # Random delay to avoid detection

        company = None
        job_title = None
        posted_time = None

        # Wait for page to load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # Try multiple selectors for job title
        title_selectors = [
            "h1.top-card-layout__title",
            "h1.topcard__title",
            "h1[class*='job-title']",
            ".job-details-jobs-unified-top-card__job-title h1",
            ".jobs-unified-top-card__job-title",
            "h1",
        ]

        for selector in title_selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element and element.text.strip():
                    job_title = element.text.strip()
                    break
            except NoSuchElementException:
                continue

        # Try multiple selectors for company name
        company_selectors = [
            "a.topcard__org-name-link",
            ".topcard__flavor a",
            "a[class*='company-name']",
            ".job-details-jobs-unified-top-card__company-name a",
            ".jobs-unified-top-card__company-name a",
            ".top-card-layout__card a[data-tracking-control-name*='company']",
            "a[href*='/company/']",
        ]

        for selector in company_selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, selector)
                if element and element.text.strip():
                    company = element.text.strip()
                    break
            except NoSuchElementException:
                continue

        # Try multiple selectors for posted time
        time_selectors = [
            ".posted-time-ago__text",
            ".topcard__flavor--metadata span",
            "span[class*='posted']",
            ".job-details-jobs-unified-top-card__primary-description-container span",
            ".jobs-unified-top-card__posted-date",
        ]

        for selector in time_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    text = element.text.strip().lower()
                    if any(word in text for word in ['ago', 'hour', 'day', 'week', 'month', 'minute']):
                        posted_time = element.text.strip()
                        break
                if posted_time:
                    break
            except NoSuchElementException:
                continue

        # Convert posted time to days
        days_ago = parse_posted_time(posted_time) if posted_time else None

        return {
            'company': company,
            'job_title': job_title,
            'days_ago': days_ago
        }

    except TimeoutException:
        print(f"Timeout loading: {url}")
        return None
    except Exception as e:
        print(f"Error extracting from {url}: {str(e)}")
        return None


def parse_posted_time(time_str):
    """Convert posted time string to number of days."""
    if not time_str:
        return None

    time_str = time_str.lower()

    # Extract number and unit
    match = re.search(r'(\d+)\s*(minute|hour|day|week|month)', time_str)
    if not match:
        return None

    num = int(match.group(1))
    unit = match.group(2)

    if 'minute' in unit:
        return 0
    elif 'hour' in unit:
        return 0
    elif 'day' in unit:
        return num
    elif 'week' in unit:
        return num * 7
    elif 'month' in unit:
        return num * 30

    return None


def main():
    excel_path = "Job Tracker.xlsx"

    # Load workbook with openpyxl to access hyperlinks
    print(f"Reading {excel_path}...")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find header row and column indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[header] = col

    print(f"Headers found: {headers}")

    # Column indices (1-based)
    link_col = headers.get('Link', 2)  # Column B
    company_col = headers.get('Company', 4)  # Column D
    job_title_col = headers.get('Job Title', 5)  # Column E
    days_col = headers.get('How long ago (Days)', 6)  # Column F

    print(f"Total rows: {ws.max_row}")

    # Find rows that need processing
    rows_to_process = []
    for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        link_cell = ws.cell(row=row, column=link_col)
        url = get_url_from_cell(link_cell)

        if is_linkedin_job_url(url):
            # Check if any of Company, Job Title, or Days columns are empty
            company = ws.cell(row=row, column=company_col).value
            job_title = ws.cell(row=row, column=job_title_col).value
            days_ago = ws.cell(row=row, column=days_col).value

            if not company or not job_title or days_ago is None:
                rows_to_process.append({
                    'row': row,
                    'url': url,
                    'missing': {
                        'company': not company,
                        'job_title': not job_title,
                        'days_ago': days_ago is None
                    }
                })

    print(f"\nFound {len(rows_to_process)} rows with LinkedIn URLs that need processing")

    if not rows_to_process:
        print("No rows to process. Exiting.")
        return

    # Display rows to be processed
    print("\nRows to process:")
    for item in rows_to_process[:10]:  # Show first 10
        print(f"  Row {item['row']}: {item['url'][:60]}...")
    if len(rows_to_process) > 10:
        print(f"  ... and {len(rows_to_process) - 10} more")

    # Set up the browser
    print("\nSetting up browser...")
    driver = setup_driver()

    try:
        # Process each row
        for i, item in enumerate(rows_to_process):
            row = item['row']
            url = item['url']
            print(f"\n[{i+1}/{len(rows_to_process)}] Processing row {row}...")
            print(f"  URL: {url[:60]}...")

            details = extract_job_details(driver, url)

            if details:
                if details['company'] and item['missing']['company']:
                    ws.cell(row=row, column=company_col).value = details['company']
                    print(f"  Company: {details['company']}")
                if details['job_title'] and item['missing']['job_title']:
                    job_cell = ws.cell(row=row, column=job_title_col)
                    job_cell.value = details['job_title']
                    # Add hyperlink to the job title cell
                    job_cell.hyperlink = url
                    # Apply blue underlined hyperlink style
                    job_cell.font = Font(color="0000FF", underline="single")
                    print(f"  Job Title: {details['job_title']} (with hyperlink)")
                if details['days_ago'] is not None and item['missing']['days_ago']:
                    ws.cell(row=row, column=days_col).value = details['days_ago']
                    print(f"  Days ago: {details['days_ago']}")

                # Save after each successful extraction (in case of crashes)
                wb.save(excel_path)
                print("  Saved to file.")
            else:
                print("  Failed to extract details")

            # Random delay between requests
            if i < len(rows_to_process) - 1:
                delay = random.uniform(3, 6)
                print(f"  Waiting {delay:.1f}s before next request...")
                time.sleep(delay)

        print(f"\nDone! Processed {len(rows_to_process)} rows.")

    finally:
        driver.quit()
        wb.save(excel_path)


if __name__ == "__main__":
    main()
